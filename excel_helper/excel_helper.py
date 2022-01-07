import re

from openpyxl import Workbook, load_workbook
from datetime import date
import calendar
import pandas
from openpyxl.cell import Cell


def connect_close_decorator(func):
    def wrap(*args, **kwargs):
        try:
            args[0].workbook = load_workbook(filename="{name}.xlsx".format(name=args[0].name), data_only=True)
        except:
            args[0].workbook.save(filename="{name}.xlsx".format(name=args[0].name))
        a = func(*args, **kwargs)
        args[0].workbook.close()
        return a

    return wrap


def connect_decorator(func):
    def wrap(*args, **kwargs):
        try:
            args[0].workbook = load_workbook(filename="{name}.xlsx".format(name=args[0].name), data_only=True)
        except:
            args[0].workbook.save(filename="{name}.xlsx".format(name=args[0].name))
        a = func(*args, **kwargs)
        return a

    return wrap


court_number = 3

weekdays_schedule = ["7-8", "8-9", "9-10", "10-11", "11-12", "12-13",
                     "13-14", "14-15", "15-16", "16-17", "17-18", "18-19",
                     "19-20", "20-21", "21-22"]
weekend_schedule = ["8-9", "9-10", "10-11", "11-12", "12-13",
                    "13-14", "14-15", "15-16", "16-17", "17-18", "18-19",
                    "19-20", "20-21"]

def weekend_check(date):
    temp_date = str(date[6:10] + "-" + date[3:5] + "-" + date[0:2])
    pandas_date = pandas.Timestamp(temp_date)
    if pandas_date.dayofweek == 6 or pandas_date.dayofweek == 5:
        return True
    return False


class Excel_helper:

    def __init__(self, name):
        self.name = name
        self.workbook = None
        self.workbook = Workbook()
        try:
            self.workbook = load_workbook(filename="{name}.xlsx".format(name=name), data_only=True)
        except:
            self.workbook.save(filename="{name}.xlsx".format(name=name))
        if len(self.workbook.worksheets) != 1:
            self.delete_worksheet("Sheet")

    # @connect_decorator
    def create_worksheet(self, name):
        try:
            t = self.workbook[name]
            return "exist"
        except:
            self.workbook.save(filename="{name}.xlsx".format(name=self.name))
            self.workbook.create_sheet(name)
            return "not_exist"

    # @connect_decorator
    def delete_worksheet(self, name):
        if name in self.workbook.worksheets:
            del self.workbook[name]

    def get_available_time(self, date):


        weekend_flag = weekend_check(date)

        if self.create_worksheet(date) != "exist":
            my_sheet = self.workbook[date]
            if not weekend_flag:
                for i in range(1, court_number + 1):
                    for j in range(1, len(weekdays_schedule) + 1):
                        if i == 1: my_sheet['A' + str(j + 1)] = weekdays_schedule[j - 1]
                        if j == 1: my_sheet[str(chr(i + 65)) + str(1)] = "court" + str(i)
                self.workbook.save(filename="{name}.xlsx".format(name=self.name))
                return weekdays_schedule
            else:
                for i in range(1, court_number + 1):
                    for j in range(1, len(weekend_schedule) + 1):
                        if i == 1: my_sheet['A' + str(j + 1)] = weekend_schedule[j - 1]
                        if j == 1: my_sheet[str(chr(i + 65)) + str(1)] = "court" + str(i)
                self.workbook.save(filename="{name}.xlsx".format(name=self.name))
                return weekend_schedule
        else:
            my_sheet = self.workbook[date]

            if not weekend_flag:
                temp_res = weekdays_schedule.copy()
                flag = False
                for j in range(2, len(weekdays_schedule) + 2):
                    flag = 0
                    for i in range(2, court_number + 2):
                        if my_sheet[str(chr(i + 64)) + str(j)].value is not None:
                            flag += 1
                    if flag == court_number:
                        temp_res.remove(weekdays_schedule[j-2])
                self.workbook.save(filename="{name}.xlsx".format(name=self.name))
                return temp_res
            if weekend_flag:
                temp_res = weekend_schedule.copy()
                flag = False
                for j in range(2, len(weekend_schedule) + 2):
                    flag = 0
                    for i in range(2, court_number + 2):
                        if my_sheet[str(chr(i + 64)) + str(j)].value is not None:
                            flag += 1
                    if flag == court_number:
                        temp_res.remove(weekend_schedule[j - 2])
                self.workbook.save(filename="{name}.xlsx".format(name=self.name))
                return temp_res

    def get_available_court(self, date, time):
        my_sheet = self.workbook[date]
        result = list()

        for i in range(2, court_number + 2):
            if my_sheet[str(chr(i + 64)) + str(weekdays_schedule.index(time)+2)].value is None:
                result.append(i-1)

        self.workbook.save(filename="{name}.xlsx".format(name=self.name))
        return result

    def get_busy_coach(self, date, time):
        my_sheet = self.workbook[date]
        coaches = list()

        for i in range(2, court_number + 2):
            if "тренер" in str(my_sheet[str(chr(i + 64)) + str(weekdays_schedule.index(time)+2)].value):
                squad = str(my_sheet[str(chr(i + 64)) + str(weekdays_schedule.index(time) + 2)].value).split()
                if len(squad) == 4:
                    coaches.append(str(squad[3]))
                if len(squad) == 5:
                    coaches.append(str(squad[3]) + " " + str(squad[4]))
        self.workbook.save(filename="{name}.xlsx".format(name=self.name))
        return coaches

    def set_property(self, date, name, time, court_number):
        my_sheet = self.workbook[date]
        my_sheet[str(chr(court_number + 65)) + str(weekdays_schedule.index(time) + 2)].value = name
        self.workbook.save(filename="{name}.xlsx".format(name=self.name))

    def set_property(self, date, name, time, court_number, coach):
        my_sheet = self.workbook[date]
        my_sheet[str(chr(court_number + 65)) + str(weekdays_schedule.index(time) + 2)].value = str(name) + ", тренер - " + str(coach)
        self.workbook.save(filename="{name}.xlsx".format(name=self.name))
